#!/usr/bin/env python3
"""
convert-workload.py — RA Toolbox (WordPress plugin): RA-CMC-hours.xlsx -> vcl-workload-hours-data.js

Reads the RA/CMC workload-hours tables from RA-CMC-hours.xlsx and regenerates
assets/js/vcl-workload-hours-data.js for the WordPress plugin from it. This
is the workload-hours counterpart to convert.py (which converts the fee
table); it follows the same conventions so both converters stay easy to
maintain side by side.

The data is wrapped as window.VCL_WORKLOAD_HD = {...} inside an IIFE (instead
of bare top-level consts), so it can never collide with same-named globals
from other plugins/the theme on a WordPress page.

Usage:
    python3 convert-workload.py path/to/RA-CMC-hours.xlsx
    python3 convert-workload.py path/to/RA-CMC-hours.xlsx -o assets/js/vcl-workload-hours-data.js

Requirement:
    pip install openpyxl

READ-ONLY: this script only ever reads the source workbook (openpyxl,
data_only=True) and never saves/writes back to it.

What this script does:
    1. Opens the workbook read-only, resolved values only (data_only=True).
    2. Reads two families of sheets, detected by their column layout rather
       than by a fixed row count, so appended rows are always picked up
       automatically:

       a) "Flat" activity sheets — one activity per row, with a single
          RA-hours or CMC-hours (min/max) pair:
            - RA - Variations & Roles       (Variation Type, Role1, Role2, RA process, min, max)
            - CMC - Variations & Roles      (Variation Type, Role1, CMC process, min, max)
            - Product Information           (same layout as RA - Variations & Roles)
            - RA - Compilation & Submission (same layout as RA - Variations & Roles)
          For CMC rows, an activeSubstance tag ("chemical" / "biological" /
          None) is derived from the process text, which encodes it as
          "(API chemical)" / "(API biological)".

       b) "Modifier" sheets — one row per (Variation Type, Role, Procedure
          Type, Active Substance) combination, with several RA-hours/
          CMC-hours (min/max) column PAIRS, one pair per "additional unit"
          dimension (e.g. "for each add. Type IA", "for each add. national").
          The dimension labels differ per sheet (Annual Update only knows
          "Type IA"; Grouping knows "Type IA/IB/II"; Super-Grouping and
          Worksharing know "national/MRP/DCP/CP"), so this script parses the
          header row of each sheet generically instead of hardcoding the
          column list — a future dimension added to the Excel is picked up
          automatically:
            - Annual Update, Grouping, Super-Grouping, Worksharing

    3. Cells containing the text "n.a." (not applicable in the source
       workbook) are converted to None rather than 0, so the browser-side
       engine can tell "no hours defined for this combination" apart from
       "defined as zero hours".
    4. Writes everything out as window.VCL_WORKLOAD_HD in
       vcl-workload-hours-data.js.
    5. Prints a summary, including a dedicated eyeball-check of the Type II
       CMC dossier-preparation hours (API chemical vs API biological) — a
       known past data-entry mistake had chemical > biological for Type II,
       which is backwards (biological dossiers require more effort). This
       script does not alter the data, only reports what it finds.
"""

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: 'openpyxl' is not installed.")
    print("Please run first:  pip install openpyxl")
    sys.exit(1)


# Sheet names as they appear in RA-CMC-hours.xlsx.
SHEET_RA_ROLES = "RA - Variations & Roles"
SHEET_CMC_ROLES = "CMC - Variations & Roles"
SHEET_PI = "Product Information"
SHEET_COMPILATION = "RA - Compilation & Submission"
SHEET_ANNUAL_UPDATE = "Annual Update"
SHEET_GROUPING = "Grouping"
SHEET_SUPER_GROUPING = "Super-Grouping"
SHEET_WORKSHARING = "Worksharing"

MODIFIER_SHEETS = [
    ("annualUpdate", SHEET_ANNUAL_UPDATE),
    ("grouping", SHEET_GROUPING),
    ("superGrouping", SHEET_SUPER_GROUPING),
    ("worksharing", SHEET_WORKSHARING),
]

# Matches process-text markers used by the CMC sheet to flag rows that are
# specific to the active substance class of the product, e.g.
# "Dossier preparation and internal check (API chemical)".
ACTIVE_SUBSTANCE_RE = re.compile(r"\(API (chemical|biological)\)", re.IGNORECASE)

# Header pattern used by the "modifier" sheets for each RA/CMC hours column,
# e.g. "RA hours (min.) for each add. Type IA [h]" or
# "CMC hours (max.) for each add. national [h]". Parsed generically so new
# dimensions (procedure types / variation types) added to the Excel in the
# future are picked up automatically without touching this script.
MODIFIER_COL_RE = re.compile(
    r"^(RA|CMC) hours \((min|max)\.\) for each add\.\s*(.+?)\s*\[h\]$"
)


def num(v):
    """Best-effort conversion of a cell value to float. The Excel file uses
    the literal string "n.a." for combinations that do not apply; those (and
    any other non-numeric text) become None rather than 0, so the browser
    engine can distinguish "not applicable" from "defined as zero hours".
    Empty cells (None) also become None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.lower() == "n.a.":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def active_substance_tag(process_text):
    """Detects the "(API chemical)" / "(API biological)" marker in a CMC
    process label and returns "chemical" / "biological" / None. The process
    text itself is left untouched (it is shown verbatim to users)."""
    if not process_text:
        return None
    m = ACTIVE_SUBSTANCE_RE.search(str(process_text))
    return m.group(1).lower() if m else None


def find_last_data_row(ws, key_col=1, first_row=2):
    """Return the last row that actually holds data, detected by a non-empty
    cell in `key_col` (1-based). Replaces a hard-coded row count so rows
    appended to the sheet are always picked up automatically."""
    last = first_row - 1
    for r in range(first_row, ws.max_row + 1):
        if ws.cell(row=r, column=key_col).value not in (None, ""):
            last = r
    return last


def load_flat_ra_sheet(ws, has_role2=True):
    """Reads a "flat" RA-hours sheet: one activity per row, columns
    Variation Type | Role1 | [Role2] | <process> | min | max. Used for
    'RA - Variations & Roles', 'Product Information' and
    'RA - Compilation & Submission' (all three share this layout; Role2 is
    present in all of them, kept as a parameter for robustness)."""
    last_row = find_last_data_row(ws)
    rows = []
    for r in range(2, last_row + 1):
        vtype = ws.cell(row=r, column=1).value
        if not vtype:
            continue
        role1 = ws.cell(row=r, column=2).value
        if has_role2:
            role2 = ws.cell(row=r, column=3).value
            process = ws.cell(row=r, column=4).value
            hmin = num(ws.cell(row=r, column=5).value)
            hmax = num(ws.cell(row=r, column=6).value)
            rows.append({
                "type": vtype, "role1": role1, "role2": role2,
                "process": process, "min": hmin, "max": hmax,
            })
        else:
            process = ws.cell(row=r, column=3).value
            hmin = num(ws.cell(row=r, column=4).value)
            hmax = num(ws.cell(row=r, column=5).value)
            rows.append({
                "type": vtype, "role1": role1,
                "process": process, "min": hmin, "max": hmax,
            })
    return rows


def load_cmc_roles_sheet(ws):
    """Reads 'CMC - Variations & Roles': Variation Type | Role1 | CMC process
    | min | max. Tags each row with activeSubstance derived from the process
    text (see active_substance_tag)."""
    last_row = find_last_data_row(ws)
    rows = []
    for r in range(2, last_row + 1):
        vtype = ws.cell(row=r, column=1).value
        if not vtype:
            continue
        process = ws.cell(row=r, column=3).value
        rows.append({
            "type": vtype,
            "role1": ws.cell(row=r, column=2).value,
            "process": process,
            "activeSubstance": active_substance_tag(process),
            "min": num(ws.cell(row=r, column=4).value),
            "max": num(ws.cell(row=r, column=5).value),
        })
    return rows


def parse_modifier_header(ws):
    """Parses the header row of a "modifier" sheet (Annual Update, Grouping,
    Super-Grouping, Worksharing). Returns the number of fixed leading columns
    (4: Variation Type, Role, Procedure Type, Active Substance) and a list of
    (column_index, stream, dimension, minmax) tuples describing every
    RA/CMC-hours column found — built from the header text itself so a
    future dimension added to the Excel (e.g. a new procedure type) is
    picked up automatically without editing this script."""
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    fixed_expected = ["Variation Type", "Role", "Procedure Type", "Active Substance"]
    if header[:4] != fixed_expected:
        print(f"  WARNING: unexpected leading columns in '{ws.title}': {header[:4]!r}")
        print(f"           expected: {fixed_expected!r}")

    cols = []
    for idx, h in enumerate(header[4:], start=5):
        if not h:
            continue
        m = MODIFIER_COL_RE.match(str(h).strip())
        if not m:
            print(f"  WARNING: unrecognised column header in '{ws.title}' col {idx}: {h!r}")
            continue
        stream_raw, minmax, dimension = m.groups()
        cols.append((idx, stream_raw.lower(), dimension, minmax))
    return cols


def load_modifier_sheet(ws):
    """Reads a "modifier" sheet generically (see parse_modifier_header):
    one row per (type, role, procedureType, activeSubstance), with nested
    ra/cmc dicts keyed by dimension label, each holding {min, max}."""
    col_defs = parse_modifier_header(ws)
    last_row = find_last_data_row(ws)
    rows = []
    for r in range(2, last_row + 1):
        vtype = ws.cell(row=r, column=1).value
        if not vtype:
            continue
        entry = {
            "type": vtype,
            "role": ws.cell(row=r, column=2).value,
            "procedureType": ws.cell(row=r, column=3).value,
            "activeSubstance": ws.cell(row=r, column=4).value,
            "ra": {},
            "cmc": {},
        }
        # Collect min/max per (stream, dimension) pair from the generically
        # detected columns.
        values = {}  # (stream, dimension) -> {"min": .., "max": ..}
        for col_idx, stream, dimension, minmax in col_defs:
            key = (stream, dimension)
            values.setdefault(key, {})[minmax] = num(ws.cell(row=r, column=col_idx).value)
        for (stream, dimension), mm in values.items():
            entry[stream][dimension] = {"min": mm.get("min"), "max": mm.get("max")}
        rows.append(entry)
    return rows


def print_type_ii_cmc_check(cmc_rows):
    """Prints the Type II CMC 'Dossier preparation and internal check' hours
    for API chemical vs API biological side by side, so a known past
    data-entry mistake (chemical entered higher than biological — backwards,
    since biological dossiers normally take longer) can be eye-checked
    against the current file. Does not alter any data."""
    print("\n  Type II CMC dossier-preparation check (API chemical vs API biological):")
    found = {"chemical": [], "biological": []}
    for row in cmc_rows:
        if row["type"] != "II":
            continue
        if row["activeSubstance"] in ("chemical", "biological") and \
           "dossier preparation" in (row["process"] or "").lower():
            found[row["activeSubstance"]].append(row)

    for row in found["chemical"]:
        print(f"    chemical   | role1={row['role1']!s:12} min={row['min']} max={row['max']}")
    for row in found["biological"]:
        print(f"    biological | role1={row['role1']!s:12} min={row['min']} max={row['max']}")

    # Compare on a common role1 basis, if possible.
    by_role_chem = {r["role1"]: r for r in found["chemical"]}
    by_role_bio = {r["role1"]: r for r in found["biological"]}
    problems = []
    for role1 in sorted(set(by_role_chem) & set(by_role_bio)):
        c, b = by_role_chem[role1], by_role_bio[role1]
        if (c["min"] or 0) > (b["min"] or 0) or (c["max"] or 0) > (b["max"] or 0):
            problems.append(role1)
    if problems:
        print(f"    WARNING: chemical hours exceed biological hours for role1={problems}"
              " — this matches the previously reported typo. Data was NOT modified.")
    elif by_role_chem:
        print("    OK: biological hours are >= chemical hours for all matched roles"
              " (no sign of the previously reported typo in this file).")
    else:
        print("    Note: could not find matching chemical/biological rows to compare.")


def main():
    parser = argparse.ArgumentParser(
        description="Converts RA-CMC-hours.xlsx into assets/js/vcl-workload-hours-data.js"
    )
    parser.add_argument("xlsx_path", type=str, help="Path to the Excel file (.xlsx)")
    parser.add_argument(
        "-o", "--output", type=str,
        default="assets/js/vcl-workload-hours-data.js",
        help="Output file (default: assets/js/vcl-workload-hours-data.js)",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.exists():
        print(f"Error: file not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading {xlsx_path} (read-only, values only) ...")
    # read_only=True guarantees this script can never accidentally write
    # back to the workbook; data_only=True resolves formulas to their last
    # calculated value.
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    print(f"  Sheets found: {wb.sheetnames}")

    expected_sheets = [
        SHEET_RA_ROLES, SHEET_CMC_ROLES, SHEET_PI, SHEET_COMPILATION,
        SHEET_ANNUAL_UPDATE, SHEET_GROUPING, SHEET_SUPER_GROUPING, SHEET_WORKSHARING,
    ]
    missing = [s for s in expected_sheets if s not in wb.sheetnames]
    if missing:
        print(f"  WARNING: expected sheet(s) not found: {missing}")

    streams = {"ra": {}, "cmc": {}}

    # --- Flat activity sheets -------------------------------------------------
    if SHEET_RA_ROLES in wb.sheetnames:
        streams["ra"][SHEET_RA_ROLES] = load_flat_ra_sheet(wb[SHEET_RA_ROLES], has_role2=True)
    if SHEET_CMC_ROLES in wb.sheetnames:
        streams["cmc"][SHEET_CMC_ROLES] = load_cmc_roles_sheet(wb[SHEET_CMC_ROLES])

    pi_rows = load_flat_ra_sheet(wb[SHEET_PI], has_role2=True) if SHEET_PI in wb.sheetnames else []
    compilation_rows = (
        load_flat_ra_sheet(wb[SHEET_COMPILATION], has_role2=True)
        if SHEET_COMPILATION in wb.sheetnames else []
    )
    streams["piActivities"] = pi_rows
    streams["compilationSubmission"] = compilation_rows

    # --- Modifier sheets (Annual Update, Grouping, Super-Grouping, Worksharing) --
    modifier_counts = {}
    for key, sheet_name in MODIFIER_SHEETS:
        if sheet_name not in wb.sheetnames:
            streams[key] = []
            continue
        rows = load_modifier_sheet(wb[sheet_name])
        streams[key] = rows
        modifier_counts[key] = len(rows)

    # --- Summary ----------------------------------------------------------
    print("\n  Row counts per stream/section:")
    print(f"    ra.{SHEET_RA_ROLES!r}: {len(streams['ra'].get(SHEET_RA_ROLES, []))}")
    print(f"    cmc.{SHEET_CMC_ROLES!r}: {len(streams['cmc'].get(SHEET_CMC_ROLES, []))}")
    print(f"    piActivities: {len(pi_rows)}")
    print(f"    compilationSubmission: {len(compilation_rows)}")
    for key, sheet_name in MODIFIER_SHEETS:
        print(f"    {key} ({sheet_name}): {len(streams.get(key, []))}")

    # --- Known-typo eyeball check: Type II CMC dossier prep, chemical vs biological
    if SHEET_CMC_ROLES in wb.sheetnames:
        print_type_ii_cmc_check(streams["cmc"][SHEET_CMC_ROLES])

    # --- Write output -------------------------------------------------------
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    data = {
        "meta": {
            "source": "RA-CMC-hours.xlsx",
            "generated": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        },
        "streams": streams,
    }

    # Wrapped in an IIFE and attached to window.VCL_WORKLOAD_HD (instead of
    # bare top-level consts) so it can never collide with same-named globals
    # from other plugins/the theme on a WordPress page.
    js = "(function(){\n"
    js += "window.VCL_WORKLOAD_HD = "
    js += json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    js += ";\n"
    js += "})();\n"
    out_path.write_text(js, encoding="utf-8")

    print(f"\nDone: wrote {out_path} ({out_path.stat().st_size:,} bytes).")
    print("\nPlease deploy the new vcl-workload-hours-data.js and spot-check a")
    print("few known type/role/process combinations in the browser before relying")
    print("on it in the additive workload engine.")


if __name__ == "__main__":
    main()
