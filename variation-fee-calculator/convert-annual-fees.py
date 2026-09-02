#!/usr/bin/env python3
"""
convert-annual-fees.py — RA Toolbox (WordPress plugin): Excel "Annual Fees" sheet
-> vcl-annual-data.js

Reads the "Annual Fees" sheet from Variation-Fee-Calculator-EU.xlsx (the same
workbook convert.py reads for the one-off variation fee table) and regenerates
assets/js/vcl-annual-data.js -- the recurring annual-maintenance-fee reference
data consumed by the Budget tool's "Plan lines — Annual maintenance fees" table.
This is the annual-fee counterpart to convert.py / convert-workload.py; it
follows the same conventions (docstring, argparse, read-only openpyxl, printed
summary) so all three converters stay easy to maintain side by side.

Usage:
    python3 convert-annual-fees.py path/to/Variation-Fee-Calculator-EU.xlsx
    python3 convert-annual-fees.py path/to/file.xlsx -o assets/js/vcl-annual-data.js

Requirement:
    pip install openpyxl

READ-ONLY: this script only ever reads the source workbook (openpyxl,
data_only=True) and never saves/writes back to it.

What this script does
----------------------
1. Reads the "Annual Fees" sheet: A=country code, B=has annual fee (yes/no),
   C=role/tariff label, D=annual fee amount (numeric or "n.a."), E=additional-
   strength fee (numeric, "n.a.", or blank -- both "n.a." and blank mean the
   fee does NOT scale with the number of strengths), F=free-text comment.
   One row per tariff variant -- most countries have exactly one row, but a
   country can have several (e.g. AT/CZ/EE/MT/NL/PL split by RMS/CMS/national,
   EU/ES/IE/UK offer several named tariffs to choose from).

2. Groups rows by country and classifies each country as:
     - no annual fee at all (column B = "no")
     - turnover-based (column B = "yes", but the row's role column is "n.a."
       with a fee also "n.a." -- the comment becomes the country's `note`,
       e.g. BE/CH/EL)
     - a flat/role-split fee, built from its row(s) into one or more
       `tariffs` entries (see step 3)

3. Turns each fee row into a tariff entry `{ id, label, role, base,
   addStrength, ccy, isDefault? }`:
     - role text that is exactly one, two, or all three of RMS/CMS/national
       (e.g. "RMS", "CMS/national", "RMS/CMS/national") is EXPLODED into one
       tariff per named role (all three combined instead collapses into a
       single role:null tariff, since the fee applies uniformly regardless
       of role) -- these never need a user pick, computeAnnualRow matches
       them by the submission's own role.
     - any other role text (a descriptive tariff name such as "Art. 10(1)/(3)
       & 10c" or "POM - standard") becomes its own role:null tariff, taken
       over VERBATIM from the sheet -- this script never paraphrases or
       shortens a label; whatever the RA/finance team writes in column C is
       what ships.
     - the country code's own currency is looked up from a small fixed
       CC_TO_CCY table (the sheet has no currency column of its own -- fee
       amounts are already in that country's local currency); everything not
       in the table is EUR.

4. Country-level `note`: every non-"default" comment found on any of a
   country's rows is collected (deduplicated) and joined into the country's
   `note` field. Nothing is ever silently dropped -- if a future row gets a
   comment, it will show up in the tool rather than vanish into the
   conversion.

5. Default-tariff safety check (HARD FAILURE, by design -- see the project's
   own decision on this): when a country ends up with more than one role:null
   tariff (a "choice" market, e.g. ES/IE/UK), the tool needs to know which one
   to price with until a user explicitly picks one. That decision is NOT
   guessed by this script. Exactly one of that country's rows must carry a
   leading "default" token in its Comments cell (case-insensitive; it may be
   followed by "; real comment text" so the marker and a genuine comment can
   coexist without either being lost -- see parse_comment). This is the same
   convention the sheet already uses for the EU "Art. 10(1)/(3) & 10c" row. If
   zero or more than one row is marked, the script prints exactly what's wrong
   and EXITS WITHOUT WRITING THE FILE -- add or fix the "default" marker in the
   Excel and re-run. This is intentional: a silently wrong default tariff would
   misprice a real client's annual fee budget.

6. FALLBACK_FX: refreshed in the same run from the "Exchange rates" sheet, for
   every non-EUR currency actually used by a generated tariff (so it never
   carries a currency the annual data doesn't need, and never misses one it
   does). Rounded to 3 significant figures, matching the file's own long-
   standing "approximate point-in-time value" convention. If the "Exchange
   rates" sheet is missing a currency this run needs, that currency is simply
   left out of FALLBACK_FX (the tool already has a visible "rate unavailable"
   status for this -- it is not a silent mispricing risk the way a wrong
   default tariff would be, so this is a warning, not a hard failure).

7. If the CURRENT assets/js/vcl-annual-data.js already exists and Node.js is
   on PATH, prints a diff against it (countries added/removed, hasAnnual /
   turnoverBased flips, and every tariff whose fee amount actually changed) --
   the eyeball-check a person should do before shipping revised fee data.

Important if the sheet structure changes in a future Excel version (new
columns, a currency column added, a different role-text convention): this
script assumes today's exact column layout (A-F, header row 1, data from row
2) and role-text vocabulary (RMS/CMS/national tokens). If the structure has
changed substantially, spot-check the printed summary against the Excel file
before deploying the regenerated vcl-annual-data.js.
"""

import argparse
import json
import re
import subprocess
import sys
from collections import OrderedDict
from math import floor, log10
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: 'openpyxl' is not installed.")
    print("Please run first:  pip install openpyxl")
    sys.exit(1)


ANNUAL_SHEET_NAME = "Annual Fees"
FX_SHEET_NAME = "Exchange rates"

# Country -> ISO currency code for the countries whose Annual Fees sheet amounts
# are in local currency, not EUR. Everything not listed here defaults to EUR.
# Kept in sync by hand with convert.py's own identical table (currency codes are
# effectively permanent, unlike fee amounts) -- if a country's currency ever
# changes, update both files.
CC_TO_CCY = {
    "CZ": "CZK", "DK": "DKK", "HU": "HUF", "IS": "ISK",
    "NO": "NOK", "PL": "PLN", "SE": "SEK", "UK": "GBP",
}

ROLE_TOKENS = ("RMS", "CMS", "national")


def num(v):
    """Excel cell -> float or None. Treats blank cells and the literal text
    "n.a." (case-insensitive) as None -- "no value defined", not zero.

    Some cells (notably the 'Exchange rates' sheet) store numbers as TEXT in
    German locale, where the comma is the DECIMAL separator (e.g. "24,19544"
    = 24.19544) and a period would be the thousands separator. Handle both
    conventions rather than blindly stripping commas (which used to corrupt
    "24,19544" into 2419544)."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s.lower() == "n.a.":
            return None
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")   # 1.234,56 -> 1234.56
        elif "," in s:
            s = s.replace(",", ".")                     # 24,19544 -> 24.19544
        try:
            return float(s)
        except ValueError:
            return None
    return float(v)


def fmt_num(n):
    """Python number -> the JS literal text for it: bare integer where exact,
    otherwise the shortest exact decimal (no trailing zeros, no scientific
    notation), matching the source file's existing style."""
    if n is None:
        return "null"
    if float(n).is_integer():
        return str(int(n))
    s = f"{n:.10f}".rstrip("0").rstrip(".")
    return s


def js_str(s):
    """Python string -> a safely-escaped JS/JSON string literal."""
    return json.dumps(s if s is not None else "", ensure_ascii=False)


def slugify(text):
    s = text.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "tariff"


def round_sig(x, sig=3):
    if x == 0:
        return 0.0
    d = sig - int(floor(log10(abs(x)))) - 1
    return round(x, d)


def parse_comment(comment):
    """Splits a Comments-column cell into (is_default, note_text). The default
    marker is a leading "default" token (case-insensitive), optionally followed
    by a ';' or ',' separator, so a row can carry BOTH the marker AND a real
    comment without either being lost:
        "default"                      -> (True,  "")
        "default; w/o Enforcement Fee" -> (True,  "w/o Enforcement Fee")
        "w/o Enforcement Fee"          -> (False, "w/o Enforcement Fee")
    """
    if not isinstance(comment, str):
        return (False, "")
    s = comment.strip()
    if not s:
        return (False, "")
    m = re.match(r"(?i)^default\b[\s;,]*", s)
    if m:
        return (True, s[m.end():].strip())
    return (False, s)


def load_annual_fee_rows(xlsx_path):
    """Reads the Annual Fees sheet into a flat list of row dicts, one per
    tariff row, preserving sheet order (already alphabetical by country)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if ANNUAL_SHEET_NAME not in wb.sheetnames:
        print(f"Error: sheet '{ANNUAL_SHEET_NAME}' not found in {xlsx_path}.")
        sys.exit(1)
    ws = wb[ANNUAL_SHEET_NAME]

    rows = []
    for r in range(2, ws.max_row + 1):
        cc = ws.cell(row=r, column=1).value
        if cc is None:
            continue
        rows.append({
            "row": r,
            "cc": str(cc).strip(),
            "has_annual": ws.cell(row=r, column=2).value,
            "role": ws.cell(row=r, column=3).value,
            "fee": ws.cell(row=r, column=4).value,
            "add_strength": ws.cell(row=r, column=5).value,
            "comment": ws.cell(row=r, column=6).value,
        })
    return rows


def load_fx_rates(xlsx_path):
    """Reads the Exchange rates sheet into a {country_code: rate} dict (1 EUR
    = X local units). Returns {} if the sheet is missing."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if FX_SHEET_NAME not in wb.sheetnames:
        return {}
    ws = wb[FX_SHEET_NAME]
    rates = {}
    for r in range(2, ws.max_row + 1):
        cc = ws.cell(row=r, column=1).value
        rate = ws.cell(row=r, column=2).value
        if cc is None or rate is None:
            continue
        rate = num(rate)
        if rate:
            rates[str(cc).strip()] = rate
    return rates


def build_countries(rows):
    """Groups the flat row list by country and builds the COUNTRIES structure.
    Returns (countries, warnings, fatal_errors). If fatal_errors is non-empty,
    the caller must NOT write the output file."""
    by_cc = OrderedDict()
    for row in rows:
        by_cc.setdefault(row["cc"], []).append(row)

    countries = []
    warnings = []
    fatal_errors = []

    for cc, cc_rows in by_cc.items():
        has_annual = str(cc_rows[0]["has_annual"]).strip().lower() == "yes"
        ccy = CC_TO_CCY.get(cc, "EUR")

        if not has_annual:
            countries.append({"cc": cc, "hasAnnual": False, "turnoverBased": False, "note": "", "tariffs": []})
            continue

        turnover_rows = [r for r in cc_rows if isinstance(r["role"], str) and r["role"].strip().lower() == "n.a."]
        other_rows = [r for r in cc_rows if r not in turnover_rows]

        if turnover_rows:
            if other_rows:
                warnings.append(
                    f"{cc}: has both a turnover-based row (row {turnover_rows[0]['row']}) and "
                    f"{len(other_rows)} priced row(s) -- only the turnover-based note was kept, "
                    "the priced row(s) were ignored. Check the sheet."
                )
            notes = []
            for r in turnover_rows:
                c = r["comment"]
                if isinstance(c, str) and c.strip():
                    notes.append(c.strip())
            countries.append({
                "cc": cc, "hasAnnual": True, "turnoverBased": True,
                "note": "; ".join(dict.fromkeys(notes)), "tariffs": [],
            })
            continue

        tariffs = []
        notes = []
        used_ids = set()

        def add_tariff(tid, label, role, base, add_strength, is_default):
            base_tid = tid
            n = 2
            while tid in used_ids:
                tid = f"{base_tid}_{n}"
                n += 1
            used_ids.add(tid)
            t = {"id": tid, "label": label, "role": role, "base": base, "addStrength": add_strength, "ccy": ccy}
            if is_default:
                t["isDefault"] = True
            tariffs.append(t)

        for r in cc_rows:
            role_text = (r["role"] or "").strip()
            fee = num(r["fee"])
            add_strength = num(r["add_strength"])
            default_flag, note_text = parse_comment(r["comment"])
            if note_text:
                notes.append(note_text)

            if fee is None:
                warnings.append(f"{cc} row {r['row']}: no numeric fee and not marked turnover-based/n.a. -- row skipped.")
                continue

            tokens = [t.strip() for t in role_text.split("/")] if role_text else []
            all_role_tokens = bool(tokens) and all(t in ROLE_TOKENS for t in tokens)

            if all_role_tokens and len(tokens) == len(ROLE_TOKENS):
                # All three roles share one uniform fee -> a single role:null tariff, labelled with
                # the combined text ("RMS/CMS/national") exactly as the sheet writes it.
                add_tariff("all", role_text, None, fee, add_strength, default_flag)
            elif all_role_tokens:
                # A subset of roles (e.g. "CMS/national" on one row) -> one tariff PER role, each
                # labelled with its own role token, not the combined text. This is the clean,
                # consistent convention computeAnnualRow matches on; it also avoids the hand-file's
                # inconsistency (CZ kept "CMS/national" while EE/MT/NL split to "CMS"/"national").
                for tok in tokens:
                    add_tariff(tok.lower(), tok, tok, fee, add_strength, default_flag)
            else:
                # A descriptive tariff name (EU legal basis, IE MA count, UK POM class, ES age) ->
                # its own role:null tariff, label taken over VERBATIM from the sheet.
                add_tariff(slugify(role_text), role_text, None, fee, add_strength, default_flag)

        choice_tariffs = [t for t in tariffs if t["role"] is None]
        if len(choice_tariffs) > 1:
            marked = [t for t in choice_tariffs if t.get("isDefault")]
            if len(marked) == 0:
                fatal_errors.append(
                    f"{cc}: {len(choice_tariffs)} tariffs need a user choice "
                    f"({', '.join(repr(t['label']) for t in choice_tariffs)}) but none is marked "
                    "'default' in the Comments column (column F). Put a leading \"default\" token in "
                    "exactly one row's comment (e.g. \"default\" or \"default; existing comment\") and re-run."
                )
            elif len(marked) > 1:
                fatal_errors.append(
                    f"{cc}: {len(marked)} rows are marked 'default' "
                    f"({', '.join(repr(t['label']) for t in marked)}) -- exactly one is required."
                )

        countries.append({
            "cc": cc, "hasAnnual": True, "turnoverBased": False,
            "note": "; ".join(dict.fromkeys(notes)), "tariffs": tariffs,
        })

    return countries, warnings, fatal_errors


def build_fallback_fx(countries, fx_rates):
    used_ccy = set()
    for c in countries:
        for t in c["tariffs"]:
            if t["ccy"] != "EUR":
                used_ccy.add(t["ccy"])

    ccy_to_cc = {v: k for k, v in CC_TO_CCY.items()}
    fx = {}
    warnings = []
    for ccy in sorted(used_ccy):
        cc = ccy_to_cc.get(ccy)
        rate = fx_rates.get(cc) if cc else None
        if rate:
            fx[ccy] = round_sig(rate, 3)
        else:
            warnings.append(f"No exchange rate found for {ccy} (country {cc}) in the '{FX_SHEET_NAME}' sheet -- FALLBACK_FX will not include it.")
    return fx, warnings


def render_js(countries, fallback_fx, generated_date):
    lines = []
    lines.append(
        "// Annual maintenance fee reference data. Generated by convert-annual-fees.py from the\n"
        '// "Annual Fees" sheet of Variation-Fee-Calculator-EU.xlsx -- DO NOT EDIT BY HAND, re-run the\n'
        "// converter instead. One entry per country; `tariffs` holds one variant per row of the sheet.\n"
        "// `role` is set only where the fee splits by RMS/CMS/national; `addStrength: null` means the\n"
        "// fee does not scale with the number of strengths. Amounts are in `ccy` units (converted to\n"
        "// EUR downstream via the shared FX rates). Dual-export like the other data modules.\n"
        "(function (root) {\n"
        '  "use strict";\n\n'
        "  var COUNTRIES = ["
    )
    for c in countries:
        head = (
            f'    {{ cc: {js_str(c["cc"])}, hasAnnual: {"true" if c["hasAnnual"] else "false"}, '
            f'turnoverBased: {"true" if c["turnoverBased"] else "false"}, note: {js_str(c["note"])}, tariffs: '
        )
        if not c["tariffs"]:
            lines.append(head + "[] },")
            continue
        lines.append(head + "[")
        for t in c["tariffs"]:
            parts = [
                f'id: {js_str(t["id"])}', f'label: {js_str(t["label"])}',
                f'role: {js_str(t["role"]) if t["role"] is not None else "null"}',
                f'base: {fmt_num(t["base"])}', f'addStrength: {fmt_num(t["addStrength"])}',
                f'ccy: {js_str(t["ccy"])}',
            ]
            if t.get("isDefault"):
                parts.append("isDefault: true")
            lines.append("      { " + ", ".join(parts) + " },")
        lines.append("    ] },")
    lines.append("  ];\n")

    lines.append(
        "  // Static fallback FX (1 EUR = X local units), used only when neither a live rate nor the\n"
        "  // calculator's own STATIC_FX_RATES cover a currency the annual dataset prices in -- see\n"
        "  // vcl-budget.js's fxByCurrency. Approximate point-in-time values, refreshed by\n"
        "  // convert-annual-fees.py from the 'Exchange rates' sheet alongside the fee data above.\n"
        "  var FALLBACK_FX = { "
        + ", ".join(f'{ccy}: {fmt_num(rate)}' for ccy, rate in fallback_fx.items())
        + " };\n"
    )
    lines.append(
        f'  var api = {{ updated: {js_str(generated_date)}, COUNTRIES: COUNTRIES, FALLBACK_FX: FALLBACK_FX }};\n'
        '  if (typeof module !== "undefined" && module.exports) module.exports = api;\n'
        "  if (root) root.VCL_ANNUAL_DATA = api;\n"
        '})(typeof window !== "undefined" ? window : this);\n'
    )
    return "\n".join(lines)


def render_json(countries, fallback_fx, generated_date):
    """The same structure as render_js(), as JSON.

    PHP cannot read the generated .js, but the fee editor has to validate what is
    typed against the shipped tariffs -- which country codes exist, which tariff
    ids a country has, and whether a tariff scales with the number of strengths.
    Written by the same run as the .js so the two cannot drift apart.
    """
    payload = {
        "updated": generated_date,
        "countries": countries,
        "fallbackFx": fallback_fx,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2) + "\n"


def load_existing_countries(output_path):
    """Best-effort: shells out to Node to require() the CURRENT output file and
    return its COUNTRIES array as plain data, for the diff report. Returns None
    if the file doesn't exist or Node isn't available -- the diff is skipped,
    never a hard failure."""
    if not output_path.exists():
        return None
    try:
        result = subprocess.run(
            ["node", "-e", f"console.log(JSON.stringify(require({json.dumps(str(output_path.resolve()))}).COUNTRIES))"],
            capture_output=True, text=True, timeout=15,
        )
        if result.returncode != 0:
            return None
        return json.loads(result.stdout)
    except Exception:
        return None


def print_diff(old_countries, new_countries):
    if old_countries is None:
        print("  (no previous vcl-annual-data.js found, or Node.js unavailable -- skipping diff)")
        return
    old_by_cc = {c["cc"]: c for c in old_countries}
    new_by_cc = {c["cc"]: c for c in new_countries}

    added = sorted(set(new_by_cc) - set(old_by_cc))
    removed = sorted(set(old_by_cc) - set(new_by_cc))
    if added:
        print(f"  + new countries: {', '.join(added)}")
    if removed:
        print(f"  - removed countries: {', '.join(removed)}")

    changed_flags = []
    changed_fees = []
    orphaned_tariffs = []
    for cc in sorted(set(old_by_cc) & set(new_by_cc)):
        old, new = old_by_cc[cc], new_by_cc[cc]
        if old.get("hasAnnual") != new.get("hasAnnual") or old.get("turnoverBased") != new.get("turnoverBased"):
            changed_flags.append(f"{cc}: hasAnnual {old.get('hasAnnual')}->{new.get('hasAnnual')}, turnoverBased {old.get('turnoverBased')}->{new.get('turnoverBased')}")
        old_t = {t["id"]: t for t in old.get("tariffs", [])}
        new_t = {t["id"]: t for t in new.get("tariffs", [])}
        for tid in sorted(set(old_t) - set(new_t)):
            changed_fees.append(f"{cc}/{tid}: removed (was {old_t[tid].get('base')})")
            orphaned_tariffs.append(f"{cc}/{tid}")
        for tid in sorted(set(new_t) - set(old_t)):
            changed_fees.append(f"{cc}/{tid}: new tariff, base={new_t[tid].get('base')}")
        for tid in sorted(set(old_t) & set(new_t)):
            if old_t[tid].get("base") != new_t[tid].get("base"):
                changed_fees.append(f"{cc}/{tid}: base {old_t[tid].get('base')} -> {new_t[tid].get('base')}")
            if old_t[tid].get("addStrength") != new_t[tid].get("addStrength"):
                changed_fees.append(f"{cc}/{tid}: addStrength {old_t[tid].get('addStrength')} -> {new_t[tid].get('addStrength')}")

    # A tariff id disappearing for a country still present (whether it was
    # dropped outright or just relabelled -- the diff can't tell those apart,
    # since the id is a slug of the Excel label) orphans any amount saved for
    # it in the fee editor's overlay: the id no longer exists to look it up by.
    for cc in sorted(set(old_by_cc) - set(new_by_cc)):
        for t in old_by_cc[cc].get("tariffs", []):
            orphaned_tariffs.append(f"{cc}/{t['id']}")

    if changed_flags:
        print("  hasAnnual/turnoverBased changes:")
        for line in changed_flags:
            print(f"    {line}")
    if changed_fees:
        print("  fee amount changes:")
        for line in changed_fees:
            print(f"    {line}")
    if orphaned_tariffs:
        print("  [!] tariff ids gone (removed or renamed in the Excel label):")
        for line in orphaned_tariffs:
            print(f"    {line}")
        print(
            "      Any annual fee saved for these ids in the fee editor's overlay "
            "(vcl_fee_overrides) is now orphaned and silently falls back to the "
            "shipped amount. Re-enter it under the new id if the tariff still exists."
        )
    if not (added or removed or changed_flags or changed_fees):
        print("  no changes vs. the current file.")


def main():
    parser = argparse.ArgumentParser(description="Converts the 'Annual Fees' sheet into assets/js/vcl-annual-data.js")
    parser.add_argument("xlsx_path", type=str, help="Path to the Excel file (.xlsx)")
    parser.add_argument("-o", "--output", type=str, default="assets/js/vcl-annual-data.js", help="Output file (default: assets/js/vcl-annual-data.js)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.exists():
        print(f"Error: file not found: {xlsx_path}")
        sys.exit(1)
    output_path = Path(args.output)

    print(f"Reading {xlsx_path} ...")
    rows = load_annual_fee_rows(xlsx_path)
    print(f"  {len(rows)} rows found in '{ANNUAL_SHEET_NAME}'.")

    countries, warnings, fatal_errors = build_countries(rows)

    if fatal_errors:
        print("\nERROR - cannot generate vcl-annual-data.js, the following must be fixed in the Excel first:")
        for e in fatal_errors:
            print(f"  [X] {e}")
        print(f"\nNothing written. {output_path} was left untouched.")
        sys.exit(1)

    if warnings:
        print("\nWarnings:")
        for w in warnings:
            print(f"  [!] {w}")

    fx_rates = load_fx_rates(xlsx_path)
    fallback_fx, fx_warnings = build_fallback_fx(countries, fx_rates)
    if fx_warnings:
        print()
        for w in fx_warnings:
            print(f"  [!] {w}")

    from datetime import date
    generated_date = date.today().isoformat()

    print(f"\nComparing against the current {output_path} ...")
    old_countries = load_existing_countries(output_path)
    print_diff(old_countries, countries)

    js = render_js(countries, fallback_fx, generated_date)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(js, encoding="utf-8")

    # The JSON is the shipped data file the fee editor validates against -- it
    # must only ever be regenerated together with the real vcl-annual-data.js,
    # never as a side effect of a preview run to some other -o path (that class
    # of bug has bitten this repo before: a converter overwrote a shipped file
    # it had no business touching).
    default_output = parser.get_default("output")
    is_default_output = args.output == default_output

    with_annual = sum(1 for c in countries if c["hasAnnual"] and not c["turnoverBased"])
    turnover = sum(1 for c in countries if c["turnoverBased"])
    no_fee = sum(1 for c in countries if not c["hasAnnual"])

    if is_default_output:
        json_path = output_path.parent.parent / "data" / "annual-fees.json"
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(render_json(countries, fallback_fx, generated_date), encoding="utf-8")
        json_line = f"    + {json_path} (structure for the fee editor)\n"
    else:
        json_line = (
            f"    (skipped: annual-fees.json is only written for the default -o "
            f"{default_output}; this was a preview run to {output_path}, so the "
            f"shipped JSON was left untouched)\n"
        )

    print(
        f"\nOK  {output_path}\n"
        f"{json_line}"
        f"    {len(countries)} countries: {with_annual} priced, {turnover} turnover-based, {no_fee} no annual fee.\n"
        f"    FALLBACK_FX: {', '.join(fallback_fx.keys()) or '(none)'}\n"
        f"    updated: {generated_date}"
    )


if __name__ == "__main__":
    main()
