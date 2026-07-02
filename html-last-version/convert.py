#!/usr/bin/env python3
"""
convert.py — Variation Fee Calculator: Excel -> data.js

Reads the fee table from the Excel file (sheet "Variation fee calculator")
and regenerates data.js for the web calculator from it.

Usage:
    python3 convert.py path/to/Variation-Fee-Calculator-EU.xlsx
    python3 convert.py path/to/file.xlsx -o data.js

Requirement:
    pip install openpyxl

What this script does:
    1. Opens the Excel file twice: once with resolved values
       (data_only=True), once with the raw formulas (data_only=False).
    2. Reads the following columns for every data row (starting at row 4):
         A  CC            (country code)
         B  Type          (IA / IB / II)
         C  Role          (RMS / CMS / national / EMA)
         D  Special cases / Comments
         E  Fee code
         F  1st strength               -> constant numeric value
         G  each add. strength         -> constant numeric value
         H  Type IA                    -> constant numeric value
         I  Type IB                    -> constant numeric value
         J  Type II                    -> constant numeric value
         K  fixed fees                 -> constant numeric value
         M,N,O,P,Q,R,S                 -> formulas (input-dependent), taken
                                          over as-is as Excel formula
                                          strings and interpreted later in
                                          the browser.
    3. Resolves known cross-sheet references (currently: the exchange-rate
       anchor 'Exchange rates'!$B$9, referenced by the Slovenian rows) and
       replaces them with the concrete numeric value, so the in-browser
       formula interpreter never needs to understand foreign-sheet syntax.
    4. Writes FEE_ROWS and COUNTRY_NAMES out as data.js.

Important if the table structure changes in a future Excel version (new
columns, shifted rows, new cross-sheet references): this script assumes
the current column order and row layout (header rows 1-3, data from row 4,
sheet name "Variation fee calculator"). If the structure has changed
substantially, please spot-check the output against the Excel file before
deploying it (the script prints a short summary at the end).
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: 'openpyxl' is not installed.")
    print("Please run first:  pip install openpyxl")
    sys.exit(1)


SHEET_NAME = "Variation fee calculator"
IMPRINT_SHEET_NAME = "Imprint"
HA_SHEET_NAME = "HA fee websites"
FIRST_DATA_ROW = 4
LAST_DATA_ROW = 419  # exclusive; adjust if more rows are added in future versions
IMPRINT_FIRST_ROW = 2
IMPRINT_LAST_ROW = 200  # exclusive; generous upper bound, stops early at first empty row

# Known cross-sheet references that the in-browser formula interpreter
# (app.js) cannot resolve on its own, because they point at a different
# sheet. If a future Excel version contains new/different such references,
# the script reports them as a warning (see resolve_sheet_refs) instead of
# silently ignoring them.
KNOWN_SHEET_REFS = {
    "'Exchange rates'!$B$9": None,  # value is looked up dynamically below
}

COUNTRY_NAMES = {
    "AT": "Austria", "BE": "Belgium", "BG": "Bulgaria", "CH": "Switzerland",
    "CY": "Cyprus", "CZ": "Czechia", "DE": "Germany", "DK": "Denmark",
    "EE": "Estonia", "EL": "Greece", "ES": "Spain", "EU": "EU EMA",
    "FI": "Finland", "FR": "France", "HR": "Croatia", "HU": "Hungary",
    "IE": "Ireland", "IS": "Iceland", "IT": "Italy", "LT": "Lithuania",
    "LU": "Luxembourg", "LV": "Latvia", "MT": "Malta", "NL": "Netherlands",
    "NO": "Norway", "PL": "Poland", "PT": "Portugal", "RO": "Romania",
    "RS": "Serbia", "SE": "Sweden", "SI": "Slovenia", "SK": "Slovakia",
    "UK": "United Kingdom",
}


# Countries whose fees are defined in local (non-EUR) currency.
# Key: ISO 3166-1 alpha-2 country code in the fee table
# Value: ISO 4217 currency code used by the Frankfurter API
CC_TO_CURRENCY = {
    "CZ": "CZK", "DK": "DKK", "HU": "HUF", "IS": "ISK",
    "NO": "NOK", "PL": "PLN", "SE": "SEK", "UK": "GBP",
    "RS": "RSD",  # RSD is NOT in the ECB/Frankfurter API — static fallback only
    "CH": "CHF",
}

def num(v):
    """Best-effort conversion of a cell value to float, keeping None as-is."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def load_rows(xlsx_path: Path):
    wb_vals = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_form = openpyxl.load_workbook(xlsx_path, data_only=False)

    if SHEET_NAME not in wb_vals.sheetnames:
        print(f"Error: sheet '{SHEET_NAME}' not found in the file.")
        print(f"Sheets present: {wb_vals.sheetnames}")
        sys.exit(1)

    ws_vals = wb_vals[SHEET_NAME]
    ws_form = wb_form[SHEET_NAME]

    # Look up the exchange-rate anchor dynamically rather than hardcoding it
    # — if the rate changes, the script automatically picks up the current
    # value from the Excel file.
    exchange_b9 = None
    if "Exchange rates" in wb_vals.sheetnames:
        exchange_b9 = wb_vals["Exchange rates"]["B9"].value
    if exchange_b9 is not None:
        KNOWN_SHEET_REFS["'Exchange rates'!$B$9"] = exchange_b9

    # Build a lookup of local-currency amounts from the "National currencies"
    # sheet (same row index, columns F-K hold the raw local amounts before
    # EUR division). We use these to support live exchange-rate conversion.
    local_amounts = {}  # row_number -> {F_lc, G_lc, H_lc, I_lc, J_lc, K_lc}
    if "National currencies" in wb_vals.sheetnames:
        ws_nc = wb_vals["National currencies"]
        for r in range(FIRST_DATA_ROW, LAST_DATA_ROW):
            cc_nc = ws_nc.cell(row=r, column=1).value
            if cc_nc and cc_nc in CC_TO_CURRENCY:
                local_amounts[r] = {
                    "F_lc": num(ws_nc.cell(row=r, column=6).value),
                    "G_lc": num(ws_nc.cell(row=r, column=7).value),
                    "H_lc": num(ws_nc.cell(row=r, column=8).value),
                    "I_lc": num(ws_nc.cell(row=r, column=9).value),
                    "J_lc": num(ws_nc.cell(row=r, column=10).value),
                    "K_lc": num(ws_nc.cell(row=r, column=11).value),
                }

    rows = []
    for r in range(FIRST_DATA_ROW, LAST_DATA_ROW):
        cc = ws_vals.cell(row=r, column=1).value
        if not cc:
            continue
        row = {
            "row": r,
            "cc": cc,
            "type": ws_vals.cell(row=r, column=2).value,
            "role": ws_vals.cell(row=r, column=3).value,
            "special": ws_vals.cell(row=r, column=4).value,
            "fee_code": ws_vals.cell(row=r, column=5).value,
            "F": num(ws_vals.cell(row=r, column=6).value),
            "G": num(ws_vals.cell(row=r, column=7).value),
            "H": num(ws_vals.cell(row=r, column=8).value),
            "I": num(ws_vals.cell(row=r, column=9).value),
            "J": num(ws_vals.cell(row=r, column=10).value),
            "K": num(ws_vals.cell(row=r, column=11).value),
            "Mf": ws_form.cell(row=r, column=13).value,
            "Nf": ws_form.cell(row=r, column=14).value,
            "Of": ws_form.cell(row=r, column=15).value,
            "Pf": ws_form.cell(row=r, column=16).value,
            "Qf": ws_form.cell(row=r, column=17).value,
            "Rf": ws_form.cell(row=r, column=18).value,
            "Sf": ws_form.cell(row=r, column=19).value,
        }
        # Attach local-currency amounts if available for live FX conversion
        if r in local_amounts:
            row["currency"] = CC_TO_CURRENCY[cc]
            row.update(local_amounts[r])
        rows.append(row)
    return rows


def load_exchange_rates(xlsx_path: Path):
    """Reads static fallback exchange rates from the 'Exchange rates' sheet.
    Returns {CC: rate} where rate = how many local units equal 1 EUR.
    These rates are used when the live API is unavailable (e.g. RS/RSD which
    the ECB does not publish)."""
    wb_vals = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Exchange rates" not in wb_vals.sheetnames:
        return {}
    ws = wb_vals["Exchange rates"]
    rates = {}
    for r in range(2, 200):
        cc = ws.cell(row=r, column=1).value
        rate = ws.cell(row=r, column=2).value
        if cc is None:
            break
        if isinstance(rate, (int, float)):
            # Map Excel CC to our country code (UK is stored as "UK", not "GB")
            rates[cc] = float(rate)
    return rates


def resolve_sheet_refs(rows):
    """Replaces known cross-sheet references in formulas with resolved
    numeric values, and reports any *unknown* cross-sheet references so
    they don't silently break the in-browser formula interpreter."""
    formula_keys = ["Mf", "Nf", "Of", "Pf", "Qf", "Rf", "Sf"]
    replaced = 0
    unknown_refs = set()

    for r in rows:
        for key in formula_keys:
            f = r.get(key)
            if not f or "!" not in f:
                continue
            # Any reference of the form 'Sheet name'!$A$1 or SheetName!A1
            sheet_refs = re.findall(r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_]*)!\$?[A-Z]+\$?\d+", f)
            for ref in sheet_refs:
                if ref in KNOWN_SHEET_REFS and KNOWN_SHEET_REFS[ref] is not None:
                    f = f.replace(ref, str(KNOWN_SHEET_REFS[ref]))
                    replaced += 1
                else:
                    unknown_refs.add((r["row"], r["cc"], key, ref))
            r[key] = f

    return replaced, unknown_refs


def build_country_names(rows):
    """Use the static English name table, but fall back gracefully to the
    raw country code if a future Excel version adds an unrecognised one —
    so the script never crashes on new countries, it just flags them."""
    codes = sorted(set(r["cc"] for r in rows))
    names = {}
    missing = []
    for c in codes:
        if c in COUNTRY_NAMES:
            names[c] = COUNTRY_NAMES[c]
        else:
            names[c] = c
            missing.append(c)
    return names, missing


def load_imprint(xlsx_path: Path):
    """Reads the changelog from the 'Imprint' sheet: column B = date,
    column C = topic/description, starting at row 2 (row 1 is the header
    'Date' / 'Topic'). Stops at the first row where both columns are
    empty. Returns a list of {date, topic} dicts in the order found in
    the sheet (the source file lists the most recent change first)."""
    wb_vals = openpyxl.load_workbook(xlsx_path, data_only=True)
    if IMPRINT_SHEET_NAME not in wb_vals.sheetnames:
        print(f"  Note: sheet '{IMPRINT_SHEET_NAME}' not found — no changelog will be included.")
        return []

    ws = wb_vals[IMPRINT_SHEET_NAME]
    entries = []
    for r in range(IMPRINT_FIRST_ROW, IMPRINT_LAST_ROW):
        date_val = ws.cell(row=r, column=2).value
        topic_val = ws.cell(row=r, column=3).value
        if date_val is None and topic_val is None:
            break
        date_str = None
        if hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%Y-%m-%d")
        elif date_val is not None:
            date_str = str(date_val)
        entries.append({"date": date_str, "topic": topic_val or ""})
    return entries


def load_ha_websites(xlsx_path: Path):
    """Reads the 'HA fee websites' sheet: column A = country code, B = link
    text + hyperlink URL, C = comments (intentionally excluded), D/E = two
    'last updated/checked' dates, F = payment method, G = annual fee flag.
    Starts at row 2 (row 1 is the header). Stops at the first row where
    column A is empty."""
    wb_vals = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_form = openpyxl.load_workbook(xlsx_path, data_only=False)
    if HA_SHEET_NAME not in wb_vals.sheetnames:
        print(f"  Note: sheet '{HA_SHEET_NAME}' not found — no HA website list will be included.")
        return []

    ws_vals = wb_vals[HA_SHEET_NAME]
    ws_form = wb_form[HA_SHEET_NAME]
    entries = []
    for r in range(2, 200):
        cc = ws_vals.cell(row=r, column=1).value
        if cc is None:
            break
        link_cell = ws_form.cell(row=r, column=2)
        link_text = link_cell.value
        link_url = link_cell.hyperlink.target if link_cell.hyperlink else None

        def fmt_date(v):
            if hasattr(v, "strftime"):
                return v.strftime("%Y-%m-%d")
            return v

        entries.append({
            "cc": cc,
            "link_text": link_text,
            "link_url": link_url,
            "updated_calc": fmt_date(ws_vals.cell(row=r, column=4).value),
            "checked_ha": fmt_date(ws_vals.cell(row=r, column=5).value),
            "payment": ws_vals.cell(row=r, column=6).value,
            "annual": ws_vals.cell(row=r, column=7).value,
        })
    return entries


def main():
    parser = argparse.ArgumentParser(description="Converts the Variation Fee Calculator Excel file into data.js")
    parser.add_argument("xlsx_path", type=str, help="Path to the Excel file (.xlsx)")
    parser.add_argument("-o", "--output", type=str, default="data.js", help="Output file (default: data.js)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.exists():
        print(f"Error: file not found: {xlsx_path}")
        sys.exit(1)

    print(f"Reading {xlsx_path} …")
    rows = load_rows(xlsx_path)
    print(f"  {len(rows)} fee rows found.")

    imprint = load_imprint(xlsx_path)
    if imprint:
        print(f"  {len(imprint)} changelog entries found (most recent: {imprint[0]['date']}).")

    ha_websites = load_ha_websites(xlsx_path)
    if ha_websites:
        print(f"  {len(ha_websites)} HA fee website entries found.")

    static_rates = load_exchange_rates(xlsx_path)
    print(f"  {len(static_rates)} static fallback exchange rate(s) loaded from Excel.")

    replaced, unknown_refs = resolve_sheet_refs(rows)
    if replaced:
        print(f"  {replaced} cross-sheet reference(s) resolved (e.g. exchange rate).")
    if unknown_refs:
        print("\n  WARNING: unknown cross-sheet formula references found:")
        for row, cc, key, ref in sorted(unknown_refs):
            print(f"    Row {row} ({cc}), column {key}: {ref}")
        print("  The web calculator currently CANNOT resolve these references.")
        print("  Please add the value manually (see KNOWN_SHEET_REFS in this script)")
        print("  or check back before deploying the new data.js.\n")

    country_names, missing_names = build_country_names(rows)
    if missing_names:
        print(f"  Note: new/unrecognised country codes without a stored display name: {missing_names}")
        print("  The country code itself will be used as the display name for now.")
        print("  Please add it to COUNTRY_NAMES in this script if you'd like a full name.\n")

    out_path = Path(args.output)
    js = "const FEE_ROWS = " + json.dumps(rows, separators=(",", ":")) + ";\n"
    js += "const COUNTRY_NAMES = " + json.dumps(country_names, ensure_ascii=False, separators=(",", ":")) + ";\n"
    js += "const IMPRINT = " + json.dumps(imprint, ensure_ascii=False, separators=(",", ":")) + ";\n"
    js += "const HA_WEBSITES = " + json.dumps(ha_websites, ensure_ascii=False, separators=(",", ":")) + ";\n"
    js += "const CC_TO_CURRENCY = " + json.dumps(CC_TO_CURRENCY, separators=(",", ":")) + ";\n"
    js += "const STATIC_FX_RATES = " + json.dumps(static_rates, separators=(",", ":")) + ";\n"
    out_path.write_text(js, encoding="utf-8")

    print(f"Done: wrote {out_path} ({out_path.stat().st_size:,} bytes).")
    print(f"Countries: {len(country_names)}, rows: {len(rows)}")
    print("\nPlease deploy the new data.js together with index.html and app.js,")
    print("and spot-check 2-3 known fees in the browser.")


if __name__ == "__main__":
    main()
